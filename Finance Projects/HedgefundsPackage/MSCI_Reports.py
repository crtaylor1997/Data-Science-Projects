import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles import Color
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os
os.chdir(r'redacted')
import datetime



def run_report(report_name, portfolio_as_of_date_str):
    """
    Runs a report set of the user's choosing that is mapped the the "processes" tab of MSCI. User needs to pass credentials via the file schema displayed below.

    :param report_name: string representing the name of an individual "process" in the processes tab of the users MSCI
    :param portfolio_as_of_date_str: string representing a date in the format "YYYYMMDD"
    :return: outputs files to the directory 
    """
    f = open(r'redacted')
    creds = f.read()
    f.close()
    creds = creds.split('\n')

    #running ths Time Series set (non wrapper)
    out = os.system(f"\redacted" prod {report_name} {portfolio_as_of_date_str} {creds[0]} {creds[1]} redacted")
    if(out==0):
        print("SUCCESS: Run {report_name}")
    else:
        print("FAILURE: Run {report_name}")

def read_msci_outputs(msci_df):
    lvl_ind = np.where(msci_df.iloc[:,1]=='Level')[0][0]
    meta_inds = np.where(msci_df.iloc[:lvl_ind,1].isna()==False)[0]
    
    meta_info = msci_df.iloc[meta_inds,1:3].values
    meta_info = dict(zip(meta_info[:,0],meta_info[:,1]))
    msci_df = msci_df.iloc[meta_inds[-1]+2:,1:]
    lvl_ind_loc = np.where(msci_df=='Level')[0][0]

    title_array = msci_df.iloc[:lvl_ind_loc+1].fillna('').values.tolist()
    column_titles = ['' for x in range(len(title_array[0]))]
    for x in range(0,len(title_array[0])):
        for y in range(0,len(title_array)):
            if(len(title_array[y][x])>0):
                if(len(column_titles[x])>0):
                    column_titles[x] = column_titles[x]+'_'+title_array[y][x]
                else:
                    column_titles[x] = column_titles[x]+title_array[y][x]

    msci_df = msci_df.loc[lvl_ind+1:]
    msci_df.columns = column_titles
    return(msci_df, meta_info)

